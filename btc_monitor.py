import requests
import openpyxl
import os
from datetime import datetime

EXCEL_FILE = "btc_history.xlsx"

def get_btc_price():
    url = "https://api.coingecko.com/api/v3/simple/price"
    params = {"ids": "bitcoin", "vs_currencies": "usd,krw"}
    res = requests.get(url, params=params).json()
    return res["bitcoin"]["usd"], res["bitcoin"]["krw"]

def get_fear_greed():
    url = "https://api.alternative.me/fng/"
    res = requests.get(url).json()
    value = res["data"][0]["value"]
    label = res["data"][0]["value_classification"]
    return int(value), label

def save_to_excel(date, btc_usd, btc_krw, fng_value, fng_label):
    if os.path.exists(EXCEL_FILE):
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "BTC History"
        headers = ["날짜", "BTC (USD)", "BTC (KRW)", "공포탐욕지수", "시장상태"]
        ws.append(headers)

        from openpyxl.styles import Font, PatternFill, Alignment
        for col in range(1, 6):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="F7931A")
            cell.alignment = Alignment(horizontal="center")

    ws.append([date, btc_usd, btc_krw, fng_value, fng_label])

    col_widths = [25, 15, 20, 15, 20]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    wb.save(EXCEL_FILE)
    print(f"✅ 엑셀 저장 완료: {EXCEL_FILE}")

def main():
    now = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
    print("📡 데이터 가져오는 중...")

    btc_usd, btc_krw = get_btc_price()
    fng_value, fng_label = get_fear_greed()

    print(f"📅 날짜: {now}")
    print(f"💰 BTC: ${btc_usd:,} USD / ₩{btc_krw:,} KRW")
    print(f"😨 공포탐욕지수: {fng_value} ({fng_label})")

    save_to_excel(now, btc_usd, btc_krw, fng_value, fng_label)

if __name__ == "__main__":
    main()